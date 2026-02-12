"""
Rutas API para reportes y dashboard
Incluye exportación a Excel con desglose de pagos adelantados y atrasados
"""
from uuid import UUID
from typing import Optional
from datetime import datetime, date, timedelta
from io import BytesIO
from fastapi import Path, APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, and_, case, extract, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from app.core.database import get_db
from app.middleware.auth import get_current_user, require_role, AuthenticatedUser
from app.models.models import (
    Transaction, TransactionStatus, Category, Unit, User, CategoryType, MonthlyReport
)

router = APIRouter(prefix="/reports", tags=["Reportes"])


# ==================== DASHBOARD ====================

@router.get("/dashboard")
async def get_dashboard(
    current_user: AuthenticatedUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Obtiene datos del dashboard principal.
    Los residentes ven solo sus datos, los admin ven todo del condominio.
    """
    # 🔍 DEBUG: Verificar condominium_id
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"🔍 Dashboard called by: {current_user.email}, Condo ID: {current_user.condominium_id}, Role: {current_user.role}")
    
    is_admin = current_user.role == "admin"
    today = date.today()
    current_month_start = today.replace(day=1)
    current_year = today.year
    current_fiscal_period = today.strftime("%Y-%m")
    
    # ✅ CORRECCIÓN CRÍTICA: Filtrar por condominio
    condominium_filter = Transaction.condominium_id == current_user.condominium_id
    
    # 🔍 1. IDENTIFICAR CATEGORÍA VIRTUAL (filtrando por condominio)
    cat_query = select(Category.id).where(
        and_(
            Category.name == "Emisión de Cuota",
            or_(
                Category.is_system == True,
                Category.condominium_id == current_user.condominium_id
            )
        )
    )
    cat_result = await db.execute(cat_query)
    internal_charge_id = cat_result.scalar()

    # Si existe la categoría, excluimos ese ID. Si no, permitimos todo (True)
    exclude_virtual = Transaction.category_id != internal_charge_id if internal_charge_id else True

    # Filtro base - Residentes ven datos del condominio completo (transparencia)
    base_filter = and_(
        Transaction.status == TransactionStatus.CONFIRMED,
        condominium_filter  # ✅ CRÍTICO
    )
    
    # Definimos qué es un gasto real (Común o Sin Categoría)
    is_common_expense_filter = or_(
        Category.is_common_expense == True,
        Category.is_common_expense == None
    )
    
    # === Totales del mes actual CON desglose de adelantados/atrasados ===
    month_query = (
        select(
            # 1. Total de ingresos
            func.coalesce(func.sum(
                case((Transaction.type == CategoryType.INCOME, Transaction.amount), else_=0)
            ), 0).label("income"),

            # 2. Gastos
            func.coalesce(func.sum(
                case((and_(
                    Transaction.type == CategoryType.EXPENSE,
                    exclude_virtual,
                    is_common_expense_filter
                ), Transaction.amount), else_=0)
            ), 0).label("expense"),

            # Ingresos adelantados
            func.coalesce(func.sum(
                case((and_(
                    Transaction.type == CategoryType.INCOME,
                    Transaction.is_advance_payment == True
                ), Transaction.amount), else_=0)
            ), 0).label("advance_income"),
            
            # Ingresos atrasados
            func.coalesce(func.sum(
                case((and_(
                    Transaction.type == CategoryType.INCOME,
                    Transaction.is_late_payment == True
                ), Transaction.amount), else_=0)
            ), 0).label("late_income")
        )
        .select_from(Transaction)
        .outerjoin(Category, Transaction.category_id == Category.id)
        .where(
            and_(
                Transaction.transaction_date >= current_month_start,
                base_filter
            )
        )
    )
    
    month_result = await db.execute(month_query)
    month_data = month_result.one()
    
    # Calcular ingresos normales
    total_income = float(month_data.income)
    advance_income = float(month_data.advance_income)
    late_income = float(month_data.late_income)
    normal_income = total_income - advance_income - late_income
    
    # === Totales del año ===
    year_start = date(current_year, 1, 1)
    year_query = (
        select(
            func.coalesce(func.sum(
                case((Transaction.type == CategoryType.INCOME, Transaction.amount), else_=0)
            ), 0).label("income"),
            
            func.coalesce(func.sum(
                case((and_(
                    Transaction.type == CategoryType.EXPENSE,
                    exclude_virtual,
                    is_common_expense_filter
                ), Transaction.amount), else_=0)
            ), 0).label("expense")
        )
        .select_from(Transaction)
        .outerjoin(Category, Transaction.category_id == Category.id)
        .where(
            and_(
                base_filter,
                Transaction.transaction_date >= year_start
            )
        )
    )
        
    year_result = await db.execute(year_query)
    year_data = year_result.one()
    
    # === Gráfica Mensual (Últimos 12 meses) ===
    twelve_months_ago = today - timedelta(days=365)
    monthly_query = select(
        extract('year', Transaction.transaction_date).label("year"),
        extract('month', Transaction.transaction_date).label("month"),
        func.sum(case((Transaction.type == "income", Transaction.amount), else_=0)).label("income"),
        func.sum(case((and_(
            Transaction.type == CategoryType.EXPENSE,
            exclude_virtual 
        ), Transaction.amount), else_=0)).label("expense")
    ).where(
        and_(
            base_filter,
            Transaction.transaction_date >= twelve_months_ago
        )
    ).group_by(
        extract('year', Transaction.transaction_date),
        extract('month', Transaction.transaction_date)
    ).order_by(
        extract('year', Transaction.transaction_date),
        extract('month', Transaction.transaction_date)
    )
    
    monthly_result = await db.execute(monthly_query)
    monthly_data = [
        {
            "year": int(row.year),
            "month": int(row.month),
            "income": float(row.income or 0),
            "expense": float(row.expense or 0),
            "balance": float((row.income or 0) - (row.expense or 0))
        }
        for row in monthly_result
    ]
    
    # === Top categorías (Pie Chart) ===
    top_categories_query = select(
        Category.name,
        Category.color,
        func.sum(Transaction.amount).label("total")
    ).join(
        Transaction, Transaction.category_id == Category.id
    ).where(
        and_(
            base_filter,
            Transaction.type == 'expense',
            exclude_virtual,
            Transaction.transaction_date >= current_month_start,
            Category.is_common_expense == True
        )
    ).group_by(
        Category.name, Category.color
    ).order_by(
        func.sum(Transaction.amount).desc()
    ).limit(5)
    
    top_cats_result = await db.execute(top_categories_query)
    top_categories = [
        {
            "name": row.name,
            "color": row.color or "#9CA3AF",
            "total": float(row.total)
        }
        for row in top_cats_result
    ]
    
    response = {
        "current_month": {
            "income": total_income,
            "expense": float(month_data.expense),
            "balance": float(total_income - float(month_data.expense)),
            "income_breakdown": {
                "normal": normal_income,
                "advance": advance_income,
                "late": late_income
            }
        },
        "current_year": {
            "income": float(year_data.income),
            "expense": float(year_data.expense),
            "balance": float(year_data.income - year_data.expense)
        },
        "monthly_trend": monthly_data,
        "top_categories": top_categories
    }
    
    # ✅ CORRECCIÓN: Estadísticas de admin filtradas por condominio
    if is_admin:
        # Estadísticas de viviendas del condominio
        units_stats = await db.execute(
            select(
                func.count(Unit.id).label("total"),
                func.count().filter(Unit.status == "OCCUPIED").label("occupied"),
                func.count().filter(Unit.balance < 0).label("with_debt"),
                func.coalesce(func.sum(Unit.balance).filter(Unit.balance < 0), 0).label("total_debt")
            ).where(Unit.condominium_id == current_user.condominium_id)  # ✅ CRÍTICO
        )
        stats_row = units_stats.one()
        
        # Transacciones pendientes del condominio
        pending_txs = await db.execute(
            select(func.count(Transaction.id))
            .where(
                and_(
                    Transaction.status == TransactionStatus.PENDING,
                    Transaction.condominium_id == current_user.condominium_id  # ✅ CRÍTICO
                )
            )
        )
        
        response["admin_stats"] = {
            "total_units": stats_row.total,
            "occupied_units": stats_row.occupied,
            "units_with_debt": stats_row.with_debt,
            "total_debt": abs(float(stats_row.total_debt)),
            "pending_transactions": pending_txs.scalar() or 0
        }
    
    return response


@router.get("/fiscal/{fiscal_period}")
async def get_fiscal_report(
    fiscal_period: str = Path(..., pattern=r'^\d{4}-(0[1-9]|1[0-2])$'),
    current_user: AuthenticatedUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Obtiene reporte fiscal de un período específico (YYYY-MM).
    """
    try:
        year, month = map(int, fiscal_period.split('-'))
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Formato de período inválido. Use YYYY-MM"
        )
    
    # ✅ CORRECCIÓN: Filtro por condominio
    base_filter = and_(
        Transaction.transaction_date >= start_date,
        Transaction.transaction_date <= end_date,
        Transaction.status == TransactionStatus.CONFIRMED,
        Transaction.condominium_id == current_user.condominium_id  # ✅ CRÍTICO
    )
    
    # Totales
    totals_query = select(
        func.sum(case((Transaction.type == "income", Transaction.amount), else_=0)).label("income"),
        func.sum(case((Transaction.type == "expense", Transaction.amount), else_=0)).label("expense")
    ).where(base_filter)
    
    totals_result = await db.execute(totals_query)
    totals = totals_result.one()
    
    # Por categoría
    by_category_query = select(
        Category.name,
        Category.type,
        func.sum(Transaction.amount).label("total")
    ).join(
        Transaction, Transaction.category_id == Category.id
    ).where(
        base_filter
    ).group_by(
        Category.name, Category.type
    ).order_by(
        Category.type, func.sum(Transaction.amount).desc()
    )
    
    categories_result = await db.execute(by_category_query)
    categories = [
        {
            "name": row.name,
            "type": row.type,
            "total": float(row.total)
        }
        for row in categories_result
    ]
    
    return {
        "fiscal_period": fiscal_period,
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "summary": {
            "income": float(totals.income or 0),
            "expense": float(totals.expense or 0),
            "balance": float((totals.income or 0) - (totals.expense or 0))
        },
        "by_category": categories
    }


@router.get("/export/unit/{unit_id}")
async def export_unit_statement(
    unit_id: UUID,
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    current_user: AuthenticatedUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Exporta estado de cuenta de una vivienda a Excel.
    """
    # ✅ CORRECCIÓN: Verificar que la unidad pertenezca al condominio
    unit_result = await db.execute(
        select(Unit)
        .options(selectinload(Unit.owner))
        .where(
            and_(
                Unit.id == unit_id,
                Unit.condominium_id == current_user.condominium_id  # ✅ CRÍTICO
            )
        )
    )
    unit = unit_result.scalar_one_or_none()
    
    if not unit:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Vivienda no encontrada en tu condominio"
        )
    
    # Verificar acceso de residentes
    if current_user.role == "resident" and str(current_user.unit_id or '') != str(unit_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="No tienes acceso a esta vivienda"
        )
    
    # Fechas por defecto
    if not end_date:
        end_date = date.today()
    if not start_date:
        start_date = end_date.replace(day=1) - timedelta(days=90)
    
    # ✅ CORRECCIÓN: Obtener transacciones DEL CONDOMINIO
    tx_query = select(Transaction).options(
        selectinload(Transaction.category)
    ).where(
        and_(
            Transaction.unit_id == unit_id,
            Transaction.transaction_date >= start_date,
            Transaction.transaction_date <= end_date,
            Transaction.status == TransactionStatus.CONFIRMED,
            Transaction.condominium_id == current_user.condominium_id  # ✅ CRÍTICO
        )
    ).order_by(Transaction.transaction_date)
    
    tx_result = await db.execute(tx_query)
    transactions = tx_result.scalars().all()
    
    # ✅ CORRECCIÓN: Saldo inicial DEL CONDOMINIO
    initial_balance_query = select(
        func.coalesce(func.sum(
            case(
                (Transaction.type == "income", Transaction.amount),
                else_=-Transaction.amount
            )
        ), 0)
    ).where(
        and_(
            Transaction.unit_id == unit_id,
            Transaction.transaction_date < start_date,
            Transaction.status == TransactionStatus.CONFIRMED,
            Transaction.condominium_id == current_user.condominium_id  # ✅ CRÍTICO
        )
    )
    
    initial_result = await db.execute(initial_balance_query)
    initial_balance = float(initial_result.scalar() or 0)
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estado de Cuenta"
    
    # Estilos
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    money_format = '#,##0.00'
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezado
    ws["A1"] = "ESTADO DE CUENTA"
    ws["A1"].font = title_font
    ws.merge_cells("A1:F1")
    
    ws["A3"] = "Unidad:"
    ws["B3"] = unit.unit_number
    ws["B3"].font = header_font
    
    ws["A4"] = "Propietario:"
    ws["B4"] = unit.owner.full_name if unit.owner else "N/A"
    
    ws["A5"] = "Período:"
    ws["B5"] = f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
    
    ws["A6"] = "Cuota mensual:"
    ws["B6"] = float(unit.monthly_fee)
    ws["B6"].number_format = money_format
    
    # Saldo inicial
    ws["A8"] = "Saldo Inicial:"
    ws["B8"] = initial_balance
    ws["B8"].number_format = money_format
    ws["A8"].font = header_font
    
    # Tabla de movimientos
    headers = ["Fecha", "Tipo", "Categoría", "Descripción", "Monto", "Saldo"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    running_balance = initial_balance
    for row_num, tx in enumerate(transactions, 11):
        if tx.type == "income":
            running_balance += float(tx.amount)
            amount = float(tx.amount)
        else:
            running_balance -= float(tx.amount)
            amount = -float(tx.amount)
        
        ws.cell(row=row_num, column=1, value=tx.transaction_date.strftime('%d/%m/%Y'))
        ws.cell(row=row_num, column=2, value="Ingreso" if tx.type == "income" else "Egreso")
        ws.cell(row=row_num, column=3, value=tx.category.name if tx.category else "")
        ws.cell(row=row_num, column=4, value=tx.description or "")
        
        amount_cell = ws.cell(row=row_num, column=5, value=amount)
        amount_cell.number_format = money_format
        
        balance_cell = ws.cell(row=row_num, column=6, value=round(running_balance, 2))
        balance_cell.number_format = money_format
        
        for col in range(1, 7):
            ws.cell(row=row_num, column=col).border = thin_border
    
    # Saldo final
    final_row = 11 + len(transactions) + 1
    ws[f"A{final_row}"] = "Saldo Final:"
    ws[f"A{final_row}"].font = header_font
    ws[f"B{final_row}"] = float(unit.balance)
    ws[f"B{final_row}"].number_format = money_format
    ws[f"B{final_row}"].font = header_font
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Guardar
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"estado_cuenta_{unit.unit_number}_{end_date.strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/debtors")
async def export_debtors_excel(
    current_user: AuthenticatedUser = Depends(require_role("admin")),
    db: AsyncSession = Depends(get_db)
):
    """
    Exporta lista de deudores a Excel.
    Solo administradores.
    """
    # ✅ CORRECCIÓN: Obtener unidades con deuda DEL CONDOMINIO
    units_query = select(Unit).options(
        selectinload(Unit.owner)
    ).where(
        and_(
            Unit.balance < 0,
            Unit.condominium_id == current_user.condominium_id  # ✅ CRÍTICO
        )
    ).order_by(Unit.balance)
    
    units_result = await db.execute(units_query)
    units = units_result.scalars().all()
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Deudores"
    
    # Estilos
    title_font = Font(bold=True, size=14)
    header_font_white = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
    money_format = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws["A1"] = "REPORTE DE DEUDORES"
    ws["A1"].font = title_font
    ws.merge_cells("A1:E1")
    
    ws["A3"] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    
    # Headers
    headers = ["Unidad", "Propietario", "Email", "Teléfono", "Saldo"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    total_debt = 0
    for row_num, unit in enumerate(units, 6):
        ws.cell(row=row_num, column=1, value=unit.unit_number)
        ws.cell(row=row_num, column=2, value=unit.owner.full_name if unit.owner else (unit.notes or "N/A"))
        ws.cell(row=row_num, column=3, value=unit.owner.email if unit.owner else "")
        ws.cell(row=row_num, column=4, value=unit.owner.phone if unit.owner else "")
        
        balance_cell = ws.cell(row=row_num, column=5, value=float(unit.balance))
        balance_cell.number_format = money_format
        
        total_debt += float(unit.balance)
        
        for col in range(1, 6):
            ws.cell(row=row_num, column=col).border = thin_border
    
    # Total
    total_row = 6 + len(units) + 1
    ws[f"D{total_row}"] = "TOTAL DEUDA:"
    ws[f"D{total_row}"].font = Font(bold=True)
    ws[f"E{total_row}"] = abs(total_debt)
    ws[f"E{total_row}"].number_format = money_format
    ws[f"E{total_row}"].font = Font(bold=True)
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Guardar
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"deudores_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
